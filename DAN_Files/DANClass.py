import pandas as pd
import xlwings as xw
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.formatting.rule import ColorScaleRule
from tqdm import tqdm
import xlsxwriter
import sys
import numpy as np
import random as rn



def num_to_col_letter(num): 
    result = ""
    while num > 0:
        num, remainder = divmod(num - 1, 26) 
        result = chr(65 + remainder) + result 
    return result


class DAN:
    def __init__(self, type, newWorkbook, excelDAN=True, MAXSUBPython=True, pythonDAN=False, allInputCategories=True, inputFormatting="clustered", orientation="horizontal", originalWorkbook=None, dataSheet=None, 
                 desiredModifications=[[]], categoryNames=True, categoryOrderPreservation=True, numericalAndAlphabeticalPreservation=False, design=True, ListOfLists=None, printStatements=True):
        self.type = type # Include desired DAN type (current supported type is "static")
        self.excelDAN = excelDAN # Make true if DAN in excel workbook is desired (excelDAN and/or pythonDAN must be selected)
        self.MAXSUBPython = MAXSUBPython # Make true if MAX/SUB count is desired in pythonDAN (this could significantly slow process if not desired)
        self.pythonDAN = pythonDAN # Make true if DAN in python is desired (excelDAN and/or pythonDAN must be selected)
        self.allInputCategories = allInputCategories # use for inputFormatting="clustered", shows all input categories in front of each input column
        self.originalWorkbook = originalWorkbook # if originalWorkbook is included, must come with dataSheet and be file path to excel book
        self.newWorkbook = newWorkbook # name of new excel book to hold the new DAN
        self.categoryOrderPreservation = categoryOrderPreservation # preserves the given category order (does not sort categories in numerical/alphabetical order)
        self.numericalAndAlphabeticalPreservation = numericalAndAlphabeticalPreservation # preserves numerical order of numerical data and alphabetical order of alphabetical data
        self.inputFormatting = inputFormatting # how the input layer is structured (current supported types are "spaced" and "clustered")
        self.dataSheet = dataSheet # if dataSheet is provided, must come with originalWorkbook and the first excel row should be column names, all rows should be same length
        self.orientation = orientation # desired orientation of DAN in excel sheet (current supported types are "horizontal"), if more than 16384 data members, program will default to vertical when supported
        self.categoryNames = categoryNames # additional formatting, include if category labels are desired
        self.desiredModifications = desiredModifications # format for this part is the following:

        ### NOTE: READ BELOW TO UNDERSTAND FORMATTING OF ABOVE LIST:

        # desiredModifications should be a LIST OF LISTS, EVEN IF ONLY ONE MODIFICATION IS DESIRED. IF NO MODIFICATIONS
        # ARE DESIRED, MAKE THIS ATTRIBUTE AN EMPTY LIST OF LISTS [[]]

        # Within each list, there will be 
        # 3 elements: 

        #### 1: The index of the column who's values are to be modified, starting at 1
        #### 2: The desired action to be done to each value in the desired column
        #### 3: The applicable specification needed (if not applicable, write None)

        # Here are the supported actions and their formatting:

        #### Splice: This action will take the range of the NUMERICAL values in the column, splice it 
        ######## into the desired number of parts, and make every value in the column one of 
        ######## these values. Should be used if values in a column are specific numbers and
        ######## overlap is desired. This will also put a ceiling on the number of unique values
        ########
        ######## Format will be [(column number), "splice", (number of splices)]
        ########
        ######## where (column number) and (number of splices) will be of type int, and "splice" will be of type string

        #### Round: This action will take every NUMERICAL value in the column and round it to a desired number of
        ######## significant figures, should also be used if there is high specificity in the values in
        ######## the column, but unique numerical values are to be preserved.
        ########
        ######## Format will be [(column number), "round", (significant figure)]
        ########
        ######## where (column number) and (significant figure) will be of type int, and "round" will be of type string


        # Multiple actions can be applied to the same column, but each action will need to be specified 
        # uniquely in the list.

        # An example desiredModifications List of Lists is the following:

        # DesiredModificationList = [[1, "round", 2], [4, "round", 6], [4, "splice", 87], [9, "round", 0]]

        # Which does the following:

        ### The values in the first column are being rounded to 2 decimal places
        ### The values in the fourth column are being rounded to 6 decimal places
        ### The values in the fourth column are being spliced and distributed into 
        ###### 87 evenly-spaced values spanning the range of the values in the column
        ### The values in the ninth column are being rounded to 0 decimal places

        self.design = design # include if formatting is desired, this may take a while for large datasets
        self.ListOfLists = ListOfLists # if ListOfLists is provided, first list should be a list of the Column Names, and all lists should be the same length, this will override the originalWorkbook and datasheet
        self.printStatements = printStatements # will give print statements to terminal on status of DAN creation if activated

    def make(self):
        ListOfListVar = True
        if self.type == "static":
            if self.originalWorkbook and self.dataSheet and self.ListOfLists:
                if self.printStatements:
                    print(" ")
                    print("Both excel path and List of List provided, default to using ListOfLists for Data")
                    print(" ")
            if self.ListOfLists:
                ListOfListVar = True
            else:
                ListOfListVar = False
            if (not self.originalWorkbook or not self.dataSheet) and not self.ListOfLists: 
                raise ValueError("Both originalWorkbook and dataSheet must be provided if not providing a ListOfLists")
            else:  
                if ListOfListVar:
                    if self.printStatements:
                        print("Initializing Data Member List")
                    self.CategoryList = self.ListOfLists[0]
                    self.DataMemberList = self.ListOfLists[1:]
                else:
                    if self.printStatements:
                        print("Initializing Data Member List")
                    wk = xw.Book(self.originalWorkbook)
                    ws1 = wk.sheets[self.dataSheet] 
                    if self.printStatements:
                        print("Initializing Category List")
                    self.CategoryList = [] 
                    y_val = 1
                    while True: 
                        cell_value = ws1.cells(1, y_val).value
                        if cell_value is None:
                            break 
                        self.CategoryList.append(cell_value)
                        y_val += 1
                    if self.printStatements:
                        print("Category List Initialized")
                        print("...")
                    testFrame = pd.read_excel(self.originalWorkbook, self.dataSheet, engine="openpyxl")
                    self.DataMemberList = testFrame.values.tolist()
                if self.printStatements:
                    print("Data Member List Initialized")
                    print("...")
                    print("Modifying Data and Initializing Dataframe")
                if self.desiredModifications[0] != []:
                    for list2 in tqdm(self.desiredModifications):
                        columnNum = list2[0] - 1
                        action = list2[1]
                        specificity = list2[2]
                        if action == "round":
                            for memberList in self.DataMemberList:
                                memberList[columnNum] = round(memberList[columnNum], specificity)
                        elif action == "splice":
                            holderList3 = []
                            for memberList1 in self.DataMemberList:
                                holderList3.append(memberList1[columnNum])
                            maxVal = max(holderList3)
                            minVal = min(holderList3)
                            range1 = maxVal - minVal
                            spliceRange = range1 / specificity
                            holderVal = minVal
                            spliceList = [minVal]
                            for i in range(specificity):
                                spliceList.append(holderVal + spliceRange)
                                holderVal += spliceRange
                            holderNum = 0
                            for memberList2 in self.DataMemberList:
                                for spliceItem in spliceList:
                                    if (memberList2[columnNum] > holderNum and memberList2[columnNum] < spliceItem) or (memberList2[columnNum] == spliceItem):
                                        memberList2[columnNum] = spliceItem
                                        break
                                    else:
                                        holderNum = spliceItem
                TheDataframe = pd.DataFrame(self.DataMemberList, columns=self.CategoryList) 
                if self.printStatements:
                    print("Data Modified and Dataframe Initialized")
                    print("...")
                    print("Initializing DAN Dataframe")
                self.BigList = [] 
                ListOfLists = []
                for category in tqdm(self.CategoryList):
                    ColumnList = TheDataframe.loc[:, category].tolist()
                    holderList1 = []
                    for thing in ColumnList:
                        if not pd.isna(thing):
                            if thing not in holderList1:
                                holderList1.append(thing)
                    if not self.numericalAndAlphabeticalPreservation:
                        holderList1.sort()
                    holderList1.insert(0, category)
                    ListOfLists.append(holderList1)
                if not self.categoryOrderPreservation:
                    NewListOfLists = sorted(ListOfLists, key=len)
                else:
                    NewListOfLists = ListOfLists
                for list1 in NewListOfLists:
                    for thing in list1:
                        self.BigList.append(thing)
                    self.BigList.append(" ")
                self.BigList.insert(0, " ")
                self.NumOfDataMembers = []
                for i in range(len(self.DataMemberList)):
                    self.NumOfDataMembers.append(i)
                if 12 + len(self.CategoryList) + len(self.DataMemberList) >= 16383 or self.orientation == "vertical":
                    self.horizontal = False
                else:
                    self.horizontal = True
                if self.printStatements:
                    print("DAN Dataframe Initialized")
                    print("...")
                    print("Filling DAN Dataframe")
                self.MaxValList = []
                InputList = []
                for item in self.BigList:
                    InputList.append(0)
                self.HolderListDFrame = [self.BigList, InputList]
                self.pythonDANList = []
                MaxValueVarDict = {}
                DANListOfLists = [self.BigList]
                rowList = []
                iterVar = 0
                if self.allInputCategories:
                    allInputVar = 3
                else:
                    allInputVar = 2
                for k, DataMem in enumerate(tqdm(self.DataMemberList)):
                    IterateList = []
                    finalString = "=ROUND(SUM("
                    rowList.append(k)
                    switch = False
                    tempDict = dict(zip(self.CategoryList, DataMem))
                    CategoryHolder = None
                    tracker = 0
                    for l, BigItem in enumerate(self.BigList):
                        if self.inputFormatting == "spaced":
                            if switch:
                                if BigItem == " ":
                                    IterateList.append(None)
                                    switch = False
                                    continue
                                elif BigItem == tempDict[CategoryHolder]:
                                    IterateList.append("={0}4".format(num_to_col_letter(14 + len(self.CategoryList) + k)))
                                    finalString += "D{0}, ".format(4 + l) 
                                else:
                                    IterateList.append(None)
                            else:
                                if BigItem in self.CategoryList:
                                    CategoryHolder = BigItem
                                    switch = True
                                    IterateList.append(None)
                                else:
                                    IterateList.append(None)
                        if self.inputFormatting == "clustered":
                            if switch:
                                if BigItem == " ":
                                    IterateList.append(None)
                                    switch = False
                                    tracker = l
                                    iterVar += 1
                                    continue
                                elif BigItem == tempDict[CategoryHolder]:
                                    IterateList.append("={0}4".format(num_to_col_letter(4 + k)))
                                    finalString += "{1}{2}{0}, ".format(4 + l - tracker, "Inputs!", num_to_col_letter(4 + allInputVar*iterVar)) 
                                else:
                                    IterateList.append(None)
                            else:
                                if BigItem in self.CategoryList:
                                    CategoryHolder = BigItem
                                    switch = True
                                    IterateList.append(None)
                                else:
                                    IterateList.append(None)
                    finalString1 = finalString[:-2]
                    finalString1 += ")/{0}, 6)".format(len(self.CategoryList))
                    IterateList[0] = finalString1
                    if self.pythonDAN:
                        MaxValueVarDict[k] = []
                        newCluster = []
                        for t, item in enumerate(IterateList[1:]):
                            if item != None:
                                MaxValueVarDict[k].append(self.HolderListDFrame[1][t + 1])
                        MaxValueVarDict[k] = round(sum(MaxValueVarDict[k])/len(self.CategoryList), 6)
                        newCluster.append(MaxValueVarDict[k])
                        for object in IterateList[1:]:
                            if object == None:
                                newCluster.append(0)
                            else:
                                newCluster.append(float(MaxValueVarDict[k]))
                        self.pythonDANList.append(newCluster)
                    DANListOfLists.append(IterateList)
                    iterVar = 0
                if not self.horizontal:
                    if self.printStatements:
                        print("wtf")
                    DANListOfLists = [list(i) for i in zip(*DANListOfLists)]
                if self.pythonDAN:
                    self.MaxValList.append(["Max Value", max(sublist[0] for sublist in self.pythonDANList)])
                prevHolder = 0
                if self.pythonDAN:
                    for p, item in enumerate(self.BigList[1:]):
                        if item in self.CategoryList and prevHolder == ' ':
                            self.MaxValList.append([item, "THISISTHEENDOFTHECATEGORY"])
                        else:
                            self.MaxValList.append([item, max(sublist[p + 1] for sublist in self.pythonDANList)])
                    prevHolder = item
                if self.printStatements:
                    print("DAN Dataframe Filled")
                    print("...")
                if self.pythonDAN:
                    if self.MAXSUBPython:
                        newBigList = ["Values"]
                        for item in self.BigList:
                            newBigList.append(item)
                        self.MAXSUBListOfListsPython = [[newBigList]]  
                        for i in range(len(self.CategoryList) + 1):
                            MAXSUBNum = round(i/len(self.CategoryList), 6)
                            holderList = [MAXSUBNum]
                            for vertical in range(len(self.BigList)):
                                if not (self.BigList[vertical] == ' ' or self.BigList[vertical] in self.CategoryList):
                                    holderVar = 0
                                    for cluster in self.pythonDANList:
                                        if cluster[vertical] == MAXSUBNum and isinstance(cluster[vertical], float):
                                            holderVar += 1
                                    holderList.append(holderVar)
                                else:
                                    holderList.append(None)
                            self.MAXSUBListOfListsPython.append(holderList)    
                    

                if self.excelDAN:
                    if self.printStatements:
                        print("Initializing and FIlling Input Dataframe")
                    InputListSpaced = []
                    TotalMaxList = []
                    for item in self.BigList:
                        InputListSpaced.append(" ")
                    if self.inputFormatting == "spaced":
                        for i, num in enumerate(tqdm(self.BigList)):
                            if (num != " ") and (num not in self.CategoryList):
                                TotalMaxList.append("=MAX({0}{1}:{2}{1})".format(num_to_col_letter(14 + len(self.CategoryList)), str(i + 4), num_to_col_letter(13 + len(self.CategoryList) + len(self.DataMemberList))))
                            else:
                                TotalMaxList.append(" ")
                        TotalMaxList[0] = "=MAX({0}4:{1}4)".format(num_to_col_letter(14 + len(self.CategoryList)), num_to_col_letter(13 + len(self.CategoryList) + len(self.DataMemberList)))
                        firstList = [self.BigList, InputListSpaced, TotalMaxList]
                    if self.inputFormatting == "clustered":
                        switch1 = True
                        holderListCluster = []
                        InputListClustered = []
                        firstList = []
                        categoryList3 = [" "]
                        for i, num in enumerate(tqdm(self.BigList)):
                            categoryList3.append(num)
                            if (num != " ") and (num not in self.CategoryList):
                                holderListCluster.append("=MAX({3}{0}{1}:{2}{1})".format("D", str(i + 4), num_to_col_letter(3 + len(self.DataMemberList)), "New_DAN!"))
                                InputListClustered.append(" ")
                            elif num in self.CategoryList:
                                holderListCluster.append(" ")
                                InputListClustered.append(" ")
                            else:
                                holderListCluster.append(" ")
                                InputListClustered.append(" ")
                                if (switch1 and i != 0 and i != 1) and not self.allInputCategories:
                                        firstList = []
                                        categoryList3[0] = ""
                                        firstList.append(categoryList3)
                                        switch1 = False
                                if (switch1 and i != 0 and i != 1) and self.allInputCategories:
                                    firstList = []
                                    switch1 = False
                                if i != 0 and i != 1 and self.allInputCategories:
                                    categoryList3[0] = ""
                                    firstList.append(categoryList3)
                                holderListCluster[0] = "=MAX({0}{1}:{0}{2})".format(num_to_col_letter(4 + len(firstList)), 6, str(3 + len(categoryList3)))
                                firstList.append(InputListClustered)
                                firstList.append(holderListCluster)
                                holderListCluster = []
                                InputListClustered = []
                                categoryList3 = []
                    if self.printStatements:
                        print("Input Dataframe Initialized and Filled")
                        print("...")
                        print("Initializing and Filling MAX VALUE Dataframe")
                    MaxCountList = []
                    for i, num in enumerate(tqdm(self.BigList)):
                        if (num != " ") and (num not in self.CategoryList):
                            MaxCountList.append(["=COUNTIF({0}{1}:{2}{1}, $E$4)".format(num_to_col_letter(14 + len(self.CategoryList)), str(i + 4), num_to_col_letter(13 + len(self.CategoryList) + len(self.DataMemberList)))])
                        else:
                            MaxCountList.append([" "])
                    if not self.horizontal:
                        MaxCountList = [list(i) for i in zip(*MaxCountList)]
                    if self.printStatements:
                        print("MAX VALUE Dataframe Initialized and Filled")
                        print("...")
                        print("Initializing and Filling MAX/SUB Count Dataframe")
                    MAXSUBListOfLists = []
                    categoryListMAXSUB = []
                    categoryListHolder = []
                    categoryListMAXSUB1 = []
                    if self.inputFormatting == "clustered":
                        for interval in tqdm(range(len(self.CategoryList) + 1)):
                            holderList2 = []
                            holderList2.append("=IF(ROUND(VALUE({0}5), 6)=ROUND(VALUE(Inputs!$B$5), 6), {1}, IF(VALUE({0}5)>VALUE(Inputs!$B$5), {3}, {2}))".format(num_to_col_letter(3 + interval), '"Max"', '"Sub"', '" "'))
                            holderList2.append("=ROUND({0}, 6)".format(interval/len(self.CategoryList)))
                            for i, item in enumerate(self.BigList):
                                if (item != " ") and (item not in self.CategoryList):
                                    holderList2.append("=COUNTIF({3}D{0}:{1}{0}, {2}5)".format(i + 4, num_to_col_letter(3 + len(self.DataMemberList)), num_to_col_letter(3 + interval), "New_DAN!"))
                                    categoryListHolder.append(" ")
                                else:
                                    holderList2.append(" ")
                                    if item in self.CategoryList:
                                        categoryListHolder.append(item)
                                        categoryListHolder.append(" ")
                            categoryListMAXSUB1.append(categoryListHolder)
                            categoryListHolder = []
                            del holderList2[2]
                            MAXSUBListOfLists.append(holderList2)
                        for element in categoryListMAXSUB1[0]:
                            categoryListMAXSUB.append([element])
                        if not self.horizontal:
                            MAXSUBListOfLists = [list(i) for i in zip(*MAXSUBListOfLists)]
                        if self.inputFormatting == "clustered":
                            if not self.allInputCategories:
                                holderList4 = [["Max Count"], ["Category"]]
                                for item in self.CategoryList:
                                    holderList4.append(["Input"])
                                    holderList4.append(["Max Value"])
                            else:
                                holderList4 = [["Max Count"]]
                                for item in range(len(self.CategoryList)):
                                    holderList4.append([NewListOfLists[item][0]])
                                    holderList4.append(["Input"])
                                    holderList4.append(["Max Value"])
                    else:
                        MAXSUBListOfLists = []
                        for interval in tqdm(range(len(self.CategoryList) + 1)):
                            holderList2 = []
                            holderList2.append("=IF(ROUND(VALUE({0}4), 6)=ROUND(VALUE($E$4), 6), {1}, IF(VALUE({0}4)>VALUE($E$4), {3}, {2}))".format(num_to_col_letter(10 + interval), '"Max"', '"Sub"', '" "'))
                            holderList2.append("=ROUND({0}, 6)".format(interval/len(self.CategoryList)))
                            for i, item in enumerate(self.BigList):
                                if (item != " ") and (item not in self.CategoryList):
                                    holderList2.append("=COUNTIF({0}{1}:{2}{1}, {3}4)".format(num_to_col_letter(14 + len(self.CategoryList)), i + 4, num_to_col_letter(13 + len(self.CategoryList) + len(self.DataMemberList)), num_to_col_letter(10 + interval)))
                                else:
                                    holderList2.append(" ")
                            del holderList2[2]
                            MAXSUBListOfLists.append(holderList2)
                    if self.printStatements:
                        print("Max/Sub Count Dataframe initialized and Filled")
                        print("...")
                    if self.inputFormatting == "spaced":
                        if self.horizontal:
                            positions = {
                                'TheOtherData': (3, 12 + len(self.CategoryList)), 
                                'FirstData': (3, 2),     
                                'SecondData': (3, 6),     
                                'MAXSUBData': (2, 9),     
                            }
                        else:
                            positions = {
                                'TheOtherData': (12 + len(self.CategoryList), 3), 
                                'FirstData': (2, 3),     
                                'SecondData': (3, 6),     
                                'MAXSUBData': (9, 2),  
                            }
                    if self.inputFormatting == "clustered":
                        positions = {
                                'TheOtherData': (3, 2), 
                                'FirstData': (4, 2),     
                                'SecondData': (4, 1), 
                                'ThirdData': (2, 1),    
                                'MAXSUBData': (3, 2),     
                            }
                    workbook = xlsxwriter.Workbook(self.newWorkbook)
                    if self.inputFormatting == "clustered":
                        inputWorksheet = workbook.add_worksheet("Inputs")
                        maxCountWorksheet = workbook.add_worksheet("MAXSUB_Counts")
                        worksheet = workbook.add_worksheet("New_DAN")
                    else:
                        worksheet = workbook.add_worksheet("New_DAN")
                    def write_list_to_sheet(data, worksheet, start_row, start_col):
                        for col_num, col_data in enumerate(data):
                            worksheet.write_column(start_row, start_col + col_num, col_data)
                    def write_list_to_sheetrow(data, worksheet, start_row, start_col):
                        for row_num, row_data in enumerate(data):
                            worksheet.write_row(start_row + row_num, start_col, row_data)
                    if self.inputFormatting == "clustered":
                        write_list_to_sheet(DANListOfLists, worksheet, *positions['TheOtherData'])
                        if self.printStatements:
                            print("Main DAN Imported")
                        write_list_to_sheet(firstList, inputWorksheet, *positions['FirstData'])
                        write_list_to_sheet(holderList4, inputWorksheet, *positions['ThirdData'])
                        if self.printStatements:
                            print("Input/MAX VALUE Imported")
                        write_list_to_sheet(MAXSUBListOfLists, maxCountWorksheet, *positions['MAXSUBData'])
                        self.BigList12 = []
                        for item in self.BigList:
                            if item in self.CategoryList:
                                self.BigList12.append([item])
                            else:
                                self.BigList12.append(list(item) if isinstance(item, (list, tuple)) else [item])
                        write_list_to_sheetrow(self.BigList12, maxCountWorksheet, *positions['SecondData'])
                        inputWorksheet.write("B5", "=MAX(E5:{0}5)".format(num_to_col_letter(2 + len(firstList))))
                        if self.printStatements:
                            print("MAX/SUB Count Imported")
                    else:
                        write_list_to_sheet(DANListOfLists, worksheet, *positions['TheOtherData'])
                        if self.printStatements:
                            print("Main DAN Imported")
                        write_list_to_sheet(firstList, worksheet, *positions['FirstData'])
                        if self.printStatements:
                            print("Input/MAX VALUE Imported")
                        write_list_to_sheet(MAXSUBListOfLists, worksheet, *positions['MAXSUBData'])
                        if self.printStatements:
                            print("MAX/SUB Count Imported")
                    if self.printStatements:
                        print("Dataframes imported to excel")
                        print("...")
                    workbook.close()
                    if self.categoryNames:
                        workbook3 = openpyxl.load_workbook(self.newWorkbook)
                        if not self.inputFormatting == "clustered":
                            worksheet3 = workbook3["New_DAN"]
                            cellA = worksheet3["{}{}".format(num_to_col_letter(13 + len(self.CategoryList)), 3)]
                            cellA.value = "Category"
                            for i in range(1, len(self.DataMemberList) + 1):
                                worksheet3["{}3".format(num_to_col_letter(13 + len(self.CategoryList) + i))].value = i
                            worksheet3["C3"].value = "Category"
                            worksheet3["D3"].value = "Input"
                            worksheet3["E3"].value = "Max Value"
                            workbook3.save(self.newWorkbook)
                        else:
                            worksheet3 = workbook3["New_DAN"]
                            cellA = worksheet3["C3"]
                            cellA.value = "Category"
                            for i in range(1, len(self.DataMemberList) + 1):
                                worksheet3["{}3".format(num_to_col_letter(3 + i))].value = i
                            workbook3.save(self.newWorkbook)
                    if self.design:
                        workbook3 = openpyxl.load_workbook(self.newWorkbook)
                        if not self.inputFormatting == "clustered":
                            worksheet3 = workbook3["New_DAN"]
                        else:
                            worksheet3 = workbook3["New_DAN"]
                            worksheet4 = workbook3["Inputs"]
                            worksheet5 = workbook3["MAXSUB_Counts"]
                        backgroundWhite = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        backgroundLightGrey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        backgroundLightBlue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                        BackgroundLighterGrey = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                        thinBorder = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
                        rightBorder = Border(left=None, right=Side(style="thin"), top=None, bottom=None)
                        rightCorner = Border(left=None, right=Side(style="thin"), top=None, bottom=Side(style="thin"))
                        underLine = Border(left=None, right=None, top=None, bottom=Side(style="thin"))
                        noBorder = Border(left=None, right=None, top=None, bottom=None)
                        LighterGreyfont = Font(color="E0E0E0")
                        if self.printStatements:
                            print("Setting Background to White")
                        if not self.inputFormatting == "clustered":
                            for row in tqdm(worksheet3.iter_rows(min_row=1, max_row=worksheet3.max_row + 1, min_col=1, max_col=worksheet3.max_column + 1)):
                                for cell in row:
                                    cell.fill = backgroundWhite
                        else:
                            for row in tqdm(worksheet3.iter_rows(min_row=1, max_row=worksheet3.max_row + 1, min_col=1, max_col=worksheet3.max_column)):
                                for cell in row:
                                    cell.fill = backgroundWhite
                            for row in tqdm(worksheet4.iter_rows(min_row=1, max_row=worksheet4.max_row, min_col=1, max_col=2 + allInputVar*len(self.CategoryList))):
                                for cell in row:
                                    cell.fill = backgroundWhite
                            for row in tqdm(worksheet5.iter_rows(min_row=1, max_row=worksheet5.max_row + 1, min_col=1, max_col=worksheet5.max_column + 2)):
                                for cell in row:
                                    cell.fill = BackgroundLighterGrey
                        if self.printStatements:
                            print("Background: White")
                            print("...")
                            print("Setting Alternate Rows to be Light Grey")
                        if not self.inputFormatting == "clustered":
                            for row in tqdm(range(1, worksheet3.max_row + 1)):
                                if worksheet3[f"C{row}"].value not in self.CategoryList:
                                    for col in range(2, worksheet3.max_column + 1):  
                                        worksheet3[f"{num_to_col_letter(col)}{row}"].fill = BackgroundLighterGrey
                        else:
                            for row in tqdm(range(2, worksheet3.max_row)):
                                if worksheet3[f"C{row}"].value not in self.CategoryList:
                                    for col in range(2, worksheet3.max_column + 1):  
                                        worksheet3[f"{num_to_col_letter(col)}{row}"].fill = BackgroundLighterGrey
                            for row in tqdm(range(3, worksheet5.max_row + 1)):
                                if worksheet5[f"C{row}"].value != " ":
                                    for col in range(2, worksheet5.max_column):  
                                        worksheet5[f"{num_to_col_letter(col)}{row}"].fill = backgroundWhite
                            for row_cells in worksheet4.iter_rows(min_row=0, max_row=4, min_col=0, max_col=2 + allInputVar*len(self.CategoryList)):
                                for cell in row_cells:   
                                    cell.fill = BackgroundLighterGrey
                        if self.printStatements:
                            print("Alternate Rows: Light Grey")
                            print("...")
                            print("Bordering Input Column")
                        if not self.inputFormatting == "clustered":
                            for row_cells in tqdm(worksheet3.iter_rows(min_row=4, max_row=worksheet3.max_row, min_col=4, max_col=4)):
                                for cell in row_cells:
                                    cell.fill = backgroundLightGrey
                                    cell.border = thinBorder
                        else:
                            for idx, row_cells in enumerate(tqdm(worksheet4.iter_cols(min_row=6, max_row=worksheet4.max_row - 1,min_col=4, max_col=2 + len(firstList)))):
                                if not self.allInputCategories:
                                    if idx % allInputVar == 0:
                                        for cell in row_cells:
                                            cell.fill = backgroundLightGrey
                                            cell.border = thinBorder
                                else:
                                    if idx % allInputVar == 0:
                                        for cell in row_cells:
                                            cell.fill = backgroundLightGrey
                                            cell.border = thinBorder
                        if self.printStatements:
                            print("Input Column: Bordered")
                            print("...")
                            print("Other Miscellaneous Formatting")
                        if not self.inputFormatting == "clustered":
                            cellA.font = Font(bold=True)
                            for row_cells in worksheet3.iter_rows(min_row=4, max_row=4, min_col=10, max_col=(10 + len(self.CategoryList))):
                                for cell in row_cells:   
                                    cell.border = thinBorder
                            for row_cells in worksheet3.iter_rows(min_row=3, max_row=3, min_col=10, max_col=(10 + len(self.CategoryList))):
                                for cell in row_cells:   
                                    cell.font = Font(bold=True)
                            for row_cells in worksheet3.iter_rows(min_row=3, max_row=3, min_col=1, max_col=worksheet3.max_column):
                                for cell in row_cells:
                                    cell.font = Font(bold=True)
                                    if cell.value and (cell.value != "Max"):
                                        cell.border = underLine
                            for row_cells in worksheet3.iter_rows(min_row=2, max_row=2, min_col=1, max_col=worksheet3.max_row):
                                for cell in row_cells:
                                    cell.font = LighterGreyfont
                            for column_cells in worksheet3.iter_rows(min_row=1, max_row=worksheet3.max_row, min_col=3, max_col=3):
                                for cell in column_cells:
                                    if cell.value in self.CategoryList:
                                        cell.font = Font(bold=True)
                        else:
                            for row_cells in worksheet3.iter_rows(min_row=3, max_row=3, min_col=3, max_col=worksheet3.max_column):
                                for cell in row_cells:   
                                    cell.font = Font(bold=True)
                                    cell.border = underLine
                            for i, row_cells in enumerate(worksheet4.iter_cols(min_row=0, max_row=worksheet4.max_row, min_col=2, max_col=2 + allInputVar*len(self.CategoryList))):
                                if self.allInputCategories:
                                    if i%allInputVar == 0:
                                        for cell in row_cells:   
                                            cell.border = rightBorder
                                else:
                                    if i%allInputVar != 0:
                                        for cell in row_cells:   
                                            cell.border = rightBorder
                            for row_cells in worksheet5.iter_rows(min_row=6, max_row=worksheet3.max_row, min_col=2, max_col=2):
                                for cell in row_cells:   
                                    cell.font = Font(bold=True)
                            for row_cells in worksheet5.iter_rows(min_row=4, max_row=4, min_col=3, max_col=worksheet3.max_column):
                                for cell in row_cells:   
                                    cell.font = Font(bold=True)
                        if self.inputFormatting == "clustered":
                            for row_cells in worksheet3.iter_rows(min_row=1, max_row=worksheet3.max_row, min_col=worksheet3.max_column, max_col=worksheet3.max_column):
                                for cell in row_cells:   
                                    cell.border = rightBorder
                            for row_cells in worksheet3.iter_rows(min_row=worksheet3.max_row, max_row=worksheet3.max_row, min_col=0, max_col=worksheet3.max_column):
                                for cell in row_cells:   
                                    cell.border = underLine
                            worksheet3["{0}{1}".format(num_to_col_letter(worksheet3.max_column), worksheet3.max_row)].border = rightCorner
                            for row_cells in worksheet4.iter_rows(min_row=worksheet4.max_row, max_row=worksheet4.max_row, min_col=0, max_col=worksheet4.max_column):
                                for cell in row_cells:   
                                    cell.border = underLine
                            worksheet4["{0}{1}".format(num_to_col_letter(worksheet4.max_column), worksheet4.max_row)].border = rightCorner
                            for row_cells in worksheet5.iter_rows(min_row=1, max_row=worksheet5.max_row, min_col=5 + len(self.CategoryList), max_col=5 + len(self.CategoryList)):
                                for cell in row_cells:   
                                    cell.border = rightBorder
                            for row_cells in worksheet5.iter_rows(min_row=worksheet5.max_row, max_row=worksheet5.max_row, min_col=0, max_col=5 + len(self.CategoryList)):
                                for cell in row_cells:   
                                    cell.border = underLine
                            worksheet5["{0}{1}".format(num_to_col_letter(len(self.CategoryList)), worksheet5.max_row)].border = rightCorner
                            for row_cells in worksheet4.iter_rows(min_row=3, max_row=3, min_col=0, max_col=5 + allInputVar*len(self.CategoryList)):
                                for cell in row_cells:   
                                    cell.font = Font(bold=True)
                            for row_cells in worksheet4.iter_rows(min_row=1, max_row=worksheet4.max_row, min_col=worksheet4.max_column, max_col=2 + allInputVar*len(self.CategoryList)):
                                for cell in row_cells:   
                                    cell.border = rightBorder
                        if not self.inputFormatting == "clustered":
                            worksheet3.conditional_formatting.add(
                            "E4:E{}".format(4 + len(self.BigList)),
                            ColorScaleRule(
                                start_type="num", start_value=0, start_color="FFFFFF",
                                end_type="num", end_value=1, end_color="69FF52"))
                            worksheet3.conditional_formatting.add(
                            "G4:G{}".format(4 + len(self.BigList)),
                            ColorScaleRule(
                                start_type="num", start_value=0, start_color="FFFFFF",
                                end_type="num", end_value=len(self.DataMemberList), end_color="69FF52"))
                            worksheet3.conditional_formatting.add(
                            "J5:{}{}".format(num_to_col_letter(len(self.CategoryList) + 10), 4 + len(self.BigList)),
                            ColorScaleRule(
                                start_type="num", start_value=0, start_color="FFFFFF",
                                end_type="num", end_value=len(self.DataMemberList)/2, end_color="69FF52"))
                            worksheet3.conditional_formatting.add(
                            "{}4:{}{}".format(num_to_col_letter(len(self.CategoryList) + 14), num_to_col_letter(len(self.CategoryList) + 14 + len(self.DataMemberList)), 4 + len(self.BigList)),
                            ColorScaleRule(
                                start_type="num", start_value=0, start_color="FFFFFF",
                                end_type="num", end_value=1, end_color="69FF52"))
                        else:
                            for i in range(5 + len(firstList)):
                                if self.allInputCategories:
                                    if i%allInputVar == 2 and i > 4:
                                        worksheet4.conditional_formatting.add(
                                        "{0}{1}:{0}{2}".format(num_to_col_letter(i), 5, worksheet4.max_row),
                                        ColorScaleRule(
                                            start_type="num", start_value=0, start_color="FFFFFF",
                                            end_type="num", end_value=1, end_color="69FF52"))
                                else:
                                    if i%allInputVar == 1 and i > 4:
                                        worksheet4.conditional_formatting.add(
                                        "{0}{1}:{0}{2}".format(num_to_col_letter(i), 5, worksheet4.max_row),
                                        ColorScaleRule(
                                            start_type="num", start_value=0, start_color="FFFFFF",
                                            end_type="num", end_value=1, end_color="69FF52"))
                                worksheet3.conditional_formatting.add(
                                "D4:{0}{1}".format(num_to_col_letter(3 + len(self.DataMemberList)), worksheet3.max_row),
                                ColorScaleRule(
                                    start_type="num", start_value=0, start_color="FFFFFF",
                                    end_type="num", end_value=1, end_color="69FF52"))
                            worksheet4.conditional_formatting.add(
                                    "B5",
                                    ColorScaleRule(
                                        start_type="num", start_value=0, start_color="FFFFFF",
                                        end_type="num", end_value=1, end_color="69FF52"))
                            worksheet5.conditional_formatting.add(
                                    "C7:{0}{1}".format(num_to_col_letter(len(self.BigList)), worksheet5.max_row),
                                    ColorScaleRule(
                                        start_type="num", start_value=0, start_color="FFFFFF",
                                        end_type="num", end_value=len(self.DataMemberList)/2, end_color="69FF52"))
                        for column_cells in worksheet3.iter_rows(min_row=1, max_row=worksheet3.max_row, min_col=13 + len(self.CategoryList), max_col=13 + len(self.CategoryList)):
                            for cell in column_cells:
                                if cell.value in self.CategoryList:
                                    cell.font = Font(bold=True)
                        if self.printStatements:
                            print("Miscellaneous formatting completed")
                            print("...")
                        workbook3.save(self.newWorkbook)
                    if self.printStatements:
                        print("Static DAN created, reload new excel book if the DAN is not present on the desired sheet :)")
        if self.type == "temporal":
            if not (self.originalWorkbook == None and self.dataSheet == None):
                print("No data provided, initializing new temporal DAN without data")
                
            else:
                if self.originalWorkbook == None or self.dataSheet == None:
                    raise ValueError("Both originalWorkbook and dataSheet must be provided")
                else:
                    print("Data provided, initializing new temporal DAN with data")
        self.InputListForPython = []


    def addInput(self, categoryInput):
        self.category = categoryInput[0]
        self.input = categoryInput[1]

        if not self.pythonDAN:
            raise ValueError("addInput is only applicable to python DANs")
        else:
            if self.printStatements:
                print(f"Adding {self.input} from {self.category}...")
            search = False
            found = False
            exist = True
            InputList = []
            for item in self.BigList:
                InputList.append(0)
            self.HolderListDFrame = [self.BigList, InputList]
            for item in self.InputListForPython:
                self.HolderListDFrame[1][item] = 1
            for t, _ in enumerate(self.HolderListDFrame[0]):
                if self.HolderListDFrame[0][t] == self.category:
                    search = True
                elif search == True and self.HolderListDFrame[0][t] == self.input and self.HolderListDFrame[1][t] == 0:
                    search = False
                    found = True
                    self.InputListForPython.append(t)
                    self.MaxValList = []
                    InputList = []
                    for item in self.BigList:
                        InputList.append(0)
                    self.HolderListDFrame = [self.BigList, InputList]
                    for item in self.InputListForPython:
                        self.HolderListDFrame[1][item] = 1
                    self.pythonDANList = []
                    MaxValueVarDict = {}
                    DANListOfLists = [self.BigList]
                    rowList = []
                    iterVar = 0
                    if self.allInputCategories:
                        allInputVar = 3
                    else:
                        allInputVar = 2
                    for k, DataMem in enumerate(tqdm(self.DataMemberList)):
                        IterateList = []
                        finalString = "=ROUND(SUM("
                        rowList.append(k)
                        switch = False
                        tempDict = dict(zip(self.CategoryList, DataMem))
                        CategoryHolder = None
                        tracker = 0
                        for l, BigItem in enumerate(self.BigList):
                            if self.inputFormatting == "spaced":
                                if switch:
                                    if BigItem == " ":
                                        IterateList.append(None)
                                        switch = False
                                        continue
                                    elif BigItem == tempDict[CategoryHolder]:
                                        IterateList.append("={0}4".format(num_to_col_letter(14 + len(self.CategoryList) + k)))
                                        finalString += "D{0}, ".format(4 + l) 
                                    else:
                                        IterateList.append(None)
                                else:
                                    if BigItem in self.CategoryList:
                                        CategoryHolder = BigItem
                                        switch = True
                                        IterateList.append(None)
                                    else:
                                        IterateList.append(None)
                            if self.inputFormatting == "clustered":
                                if switch:
                                    if BigItem == " ":
                                        IterateList.append(None)
                                        switch = False
                                        tracker = l
                                        iterVar += 1
                                        continue
                                    elif BigItem == tempDict[CategoryHolder]:
                                        IterateList.append("={0}4".format(num_to_col_letter(4 + k)))
                                        finalString += "{1}{2}{0}, ".format(4 + l - tracker, "Inputs!", num_to_col_letter(4 + allInputVar*iterVar)) 
                                    else:
                                        IterateList.append(None)
                                else:
                                    if BigItem in self.CategoryList:
                                        CategoryHolder = BigItem
                                        switch = True
                                        IterateList.append(None)
                                    else:
                                        IterateList.append(None)
                        finalString1 = finalString[:-2]
                        finalString1 += ")/{0}, 6)".format(len(self.CategoryList))
                        IterateList[0] = finalString1
                        if self.pythonDAN:
                            MaxValueVarDict[k] = []
                            newCluster = []
                            for t, item in enumerate(IterateList[1:]):
                                if item != None:
                                    MaxValueVarDict[k].append(self.HolderListDFrame[1][t + 1])
                            MaxValueVarDict[k] = round(sum(MaxValueVarDict[k])/len(self.CategoryList), 6)
                            newCluster.append(MaxValueVarDict[k])
                            for object in IterateList[1:]:
                                if object == None:
                                    newCluster.append(0)
                                else:
                                    newCluster.append(MaxValueVarDict[k])
                            self.pythonDANList.append(newCluster)
                        DANListOfLists.append(IterateList)
                        iterVar = 0
                    if not self.horizontal:
                        print("wtf")
                        DANListOfLists = [list(i) for i in zip(*DANListOfLists)]
                    self.MaxValList.append(["Max Value", max(sublist[0] for sublist in self.pythonDANList)])
                    prevHolder = 0
                    for p, item in enumerate(self.BigList[1:]):
                        if item in self.CategoryList and prevHolder == ' ':
                            self.MaxValList.append([item, "THISISTHEENDOFTHECATEGORY"])
                        else:
                            self.MaxValList.append([item, max(sublist[p + 1] for sublist in self.pythonDANList)])
                        prevHolder = item 
                    if self.MAXSUBPython:
                        newBigList = ["Values"]
                        for item in self.BigList:
                            newBigList.append(item)
                        self.MAXSUBListOfListsPython = [[newBigList]]  
                        for i in tqdm(range(len(self.CategoryList) + 1)):
                            MAXSUBNum = round(i/len(self.CategoryList), 6)
                            holderList = [MAXSUBNum]
                            for vertical in range(len(self.BigList)):
                                if not (self.BigList[vertical] == ' ' or self.BigList[vertical] in self.CategoryList):
                                    holderVar = 0
                                    for cluster in self.pythonDANList:
                                        if cluster[vertical] == MAXSUBNum and isinstance(cluster[vertical], float):
                                            holderVar += 1
                                    holderList.append(holderVar)
                                else:
                                    holderList.append(None)
                            self.MAXSUBListOfListsPython.append(holderList)   
                    break
                elif search == True and self.HolderListDFrame[0][t] == self.input and self.HolderListDFrame[1][t] == 1:
                    if self.printStatements:
                        print("This input is already activated")
                    exist = False
                elif search == True and self.HolderListDFrame[0][t] in self.CategoryList and self.HolderListDFrame[0][t - 1] == ' ':
                    if self.printStatements:
                        print(f"Either the input {self.input} or the category {self.category} does not exist")  
                    exist = False
                    break
            if found and exist:
                if self.printStatements:
                    print(f"{self.input} from {self.category} has been Added!")  
                    print(" ") 
            

    def removeInput(self, categoryInput):
        self.category = categoryInput[0]
        self.input = categoryInput[1]

        if not self.pythonDAN:
            raise ValueError("removeInput is only applicable to python DANs")
        else:
            if self.printStatements:
                print(f"Removing {self.input} from {self.category}...")
            search = False
            found = False
            exist = True
            self.MaxValList = []
            InputList = []
            for item in self.BigList:
                InputList.append(0)
            self.HolderListDFrame = [self.BigList, InputList]
            for item in self.InputListForPython:
                self.HolderListDFrame[1][item] = 1
            for t, _ in enumerate(self.HolderListDFrame[0]):
                if self.HolderListDFrame[0][t] == self.category:
                    search = True
                elif search == True and self.HolderListDFrame[0][t] == self.input and self.HolderListDFrame[1][t] == 1:
                    self.HolderListDFrame[1][t] = 0
                    search = False
                    found = True
                    self.pythonDANList = []
                    MaxValueVarDict = {}
                    DANListOfLists = [self.BigList]
                    rowList = []
                    iterVar = 0
                    for index, value in enumerate(self.InputListForPython):
                        if value == t:
                            self.InputListForPython.pop(index)
                    if self.allInputCategories:
                        allInputVar = 3
                    else:
                        allInputVar = 2
                    for k, DataMem in enumerate(tqdm(self.DataMemberList)):
                        IterateList = []
                        finalString = "=ROUND(SUM("
                        rowList.append(k)
                        switch = False
                        tempDict = dict(zip(self.CategoryList, DataMem))
                        CategoryHolder = None
                        tracker = 0
                        for l, BigItem in enumerate(self.BigList):
                            if self.inputFormatting == "spaced":
                                if switch:
                                    if BigItem == " ":
                                        IterateList.append(None)
                                        switch = False
                                        continue
                                    elif BigItem == tempDict[CategoryHolder]:
                                        IterateList.append("={0}4".format(num_to_col_letter(14 + len(self.CategoryList) + k)))
                                        finalString += "D{0}, ".format(4 + l) 
                                    else:
                                        IterateList.append(None)
                                else:
                                    if BigItem in self.CategoryList:
                                        CategoryHolder = BigItem
                                        switch = True
                                        IterateList.append(None)
                                    else:
                                        IterateList.append(None)
                            if self.inputFormatting == "clustered":
                                if switch:
                                    if BigItem == " ":
                                        IterateList.append(None)
                                        switch = False
                                        tracker = l
                                        iterVar += 1
                                        continue
                                    elif BigItem == tempDict[CategoryHolder]:
                                        IterateList.append("={0}4".format(num_to_col_letter(4 + k)))
                                        finalString += "{1}{2}{0}, ".format(4 + l - tracker, "Inputs!", num_to_col_letter(4 + allInputVar*iterVar)) 
                                    else:
                                        IterateList.append(None)
                                else:
                                    if BigItem in self.CategoryList:
                                        CategoryHolder = BigItem
                                        switch = True
                                        IterateList.append(None)
                                    else:
                                        IterateList.append(None)
                        finalString1 = finalString[:-2]
                        finalString1 += ")/{0}, 6)".format(len(self.CategoryList))
                        IterateList[0] = finalString1
                        if self.pythonDAN:
                            MaxValueVarDict[k] = []
                            newCluster = []
                            for t, item in enumerate(IterateList[1:]):
                                if item != None:
                                    MaxValueVarDict[k].append(self.HolderListDFrame[1][t + 1])
                            MaxValueVarDict[k] = round(sum(MaxValueVarDict[k])/len(self.CategoryList), 6)
                            newCluster.append(MaxValueVarDict[k])
                            for object in IterateList[1:]:
                                if object == None:
                                    newCluster.append(0)
                                else:
                                    newCluster.append(MaxValueVarDict[k])
                            self.pythonDANList.append(newCluster)
                        DANListOfLists.append(IterateList)
                        iterVar = 0
                    if not self.horizontal:
                        print("wtf")
                        DANListOfLists = [list(i) for i in zip(*DANListOfLists)]
                    self.MaxValList.append(["Max Value", max(sublist[0] for sublist in self.pythonDANList)])
                    prevHolder = 0
                    for p, item in enumerate(self.BigList[1:]):
                        if item in self.CategoryList and prevHolder == ' ':
                            self.MaxValList.append([item, "THISISTHEENDOFTHECATEGORY"])
                        else:
                            self.MaxValList.append([item, max(sublist[p + 1] for sublist in self.pythonDANList)])
                        prevHolder = item  
                    if self.MAXSUBPython:
                        newBigList = ["Values"]
                        for item in self.BigList:
                            newBigList.append(item)
                        self.MAXSUBListOfListsPython = [[newBigList]]  
                        for i in tqdm(range(len(self.CategoryList) + 1)):
                            MAXSUBNum = round(i/len(self.CategoryList), 6)
                            holderList = [MAXSUBNum]
                            for vertical in range(len(self.BigList)):
                                if not (self.BigList[vertical] == ' ' or self.BigList[vertical] in self.CategoryList):
                                    holderVar = 0
                                    for cluster in self.pythonDANList:
                                        if cluster[vertical] == MAXSUBNum and isinstance(cluster[vertical], float):
                                            holderVar += 1
                                    holderList.append(holderVar)
                                else:
                                    holderList.append(None)
                            self.MAXSUBListOfListsPython.append(holderList)  
                    break
                elif search == True and self.HolderListDFrame[0][t] == self.input and self.HolderListDFrame[1][t] == 0:
                    if self.printStatements:
                        print("This input is not currently activated")
                    exist = False
                elif search == True and self.HolderListDFrame[0][t] in self.CategoryList and self.HolderListDFrame[0][t - 1] == ' ':
                    if self.printStatements:
                        print(f"Either the input {self.input} or the category {self.category} does not exist")  
                    exist = False
                    break
            if found and exist:
                if self.printStatements:
                    print(f"{self.input} from {self.category} has been Removed!")  
                    print(" ") 
            
    def replaceInputsWith(self, newInputsList):
        self.newInputsList = newInputsList

        if not self.pythonDAN:
            raise ValueError("addInput is only applicable to python DANs")
        else:
            if self.printStatements:
                print(f"Replacing Inputs...")
            search = False
            InputList = []
            for item in self.BigList:
                InputList.append(0)
            
            self.HolderListDFrame = [self.BigList, InputList]
            self.InputListForPython = []
            for element in newInputsList:
                category = element[0]
                input = element[1]
                for t, _ in enumerate(self.HolderListDFrame[0]):
                    if self.HolderListDFrame[0][t] == category:
                        search = True
                    elif search == True and self.HolderListDFrame[0][t] == input and self.HolderListDFrame[1][t] == 0:
                        search = False
                        self.InputListForPython.append(t)
            InputList = []
            for item in self.BigList:
                InputList.append(0)
            self.HolderListDFrame = [self.BigList, InputList]
            for item in self.InputListForPython:
                self.HolderListDFrame[1][item] = 1
            for t, _ in enumerate(self.HolderListDFrame[0]):
                if self.HolderListDFrame[1][t] == 1:
                    self.MaxValList = []
                    InputList = []
                    self.pythonDANList = []
                    MaxValueVarDict = {}
                    DANListOfLists = [self.BigList]
                    rowList = []
                    iterVar = 0
                    if self.allInputCategories:
                        allInputVar = 3
                    else:
                        allInputVar = 2
                    for k, DataMem in enumerate(tqdm(self.DataMemberList)):
                        IterateList = []
                        finalString = "=ROUND(SUM("
                        rowList.append(k)
                        switch = False
                        tempDict = dict(zip(self.CategoryList, DataMem))
                        CategoryHolder = None
                        tracker = 0
                        for l, BigItem in enumerate(self.BigList):
                            if self.inputFormatting == "spaced":
                                if switch:
                                    if BigItem == " ":
                                        IterateList.append(None)
                                        switch = False
                                        continue
                                    elif BigItem == tempDict[CategoryHolder]:
                                        IterateList.append("={0}4".format(num_to_col_letter(14 + len(self.CategoryList) + k)))
                                        finalString += "D{0}, ".format(4 + l) 
                                    else:
                                        IterateList.append(None)
                                else:
                                    if BigItem in self.CategoryList:
                                        CategoryHolder = BigItem
                                        switch = True
                                        IterateList.append(None)
                                    else:
                                        IterateList.append(None)
                            if self.inputFormatting == "clustered":
                                if switch:
                                    if BigItem == " ":
                                        IterateList.append(None)
                                        switch = False
                                        tracker = l
                                        iterVar += 1
                                        continue
                                    elif BigItem == tempDict[CategoryHolder]:
                                        IterateList.append("={0}4".format(num_to_col_letter(4 + k)))
                                        finalString += "{1}{2}{0}, ".format(4 + l - tracker, "Inputs!", num_to_col_letter(4 + allInputVar*iterVar)) 
                                    else:
                                        IterateList.append(None)
                                else:
                                    if BigItem in self.CategoryList:
                                        CategoryHolder = BigItem
                                        switch = True
                                        IterateList.append(None)
                                    else:
                                        IterateList.append(None)
                        finalString1 = finalString[:-2]
                        finalString1 += ")/{0}, 6)".format(len(self.CategoryList))
                        IterateList[0] = finalString1
                        if self.pythonDAN:
                            MaxValueVarDict[k] = []
                            newCluster = []
                            for t, item in enumerate(IterateList[1:]):
                                if item != None:
                                    MaxValueVarDict[k].append(self.HolderListDFrame[1][t + 1])
                            MaxValueVarDict[k] = round(sum(MaxValueVarDict[k])/len(self.CategoryList), 6)
                            newCluster.append(MaxValueVarDict[k])
                            for object in IterateList[1:]:
                                if object == None:
                                    newCluster.append(0)
                                else:
                                    newCluster.append(MaxValueVarDict[k])
                            self.pythonDANList.append(newCluster)
                        DANListOfLists.append(IterateList)
                        iterVar = 0
                    if not self.horizontal:
                        print("wtf")
                        DANListOfLists = [list(i) for i in zip(*DANListOfLists)]
                    self.MaxValList.append(["Max Value", max(sublist[0] for sublist in self.pythonDANList)])
                    prevHolder = 0
                    for p, item in enumerate(self.BigList[1:]):
                        if item in self.CategoryList and prevHolder == ' ':
                            self.MaxValList.append([item, "THISISTHEENDOFTHECATEGORY"])
                        else:
                            self.MaxValList.append([item, max(sublist[p + 1] for sublist in self.pythonDANList)])
                        prevHolder = item 
                    if self.MAXSUBPython:
                        newBigList = ["Values"]
                        for item in self.BigList:
                            newBigList.append(item)
                        self.MAXSUBListOfListsPython = [[newBigList]]  
                        for i in tqdm(range(len(self.CategoryList) + 1)):
                            MAXSUBNum = round(i/len(self.CategoryList), 6)
                            holderList = [MAXSUBNum]
                            for vertical in range(len(self.BigList)):
                                if not (self.BigList[vertical] == ' ' or self.BigList[vertical] in self.CategoryList):
                                    holderVar = 0
                                    for cluster in self.pythonDANList:
                                        if cluster[vertical] == MAXSUBNum and isinstance(cluster[vertical], float):
                                            holderVar += 1
                                    holderList.append(holderVar)
                                else:
                                    holderList.append(None)
                            self.MAXSUBListOfListsPython.append(holderList)   
                    break
            if self.printStatements:
                print(f"Inputs have been Replaced!")  
                print(" ") 

    def showClusters(self):
        print("**************************")
        print("Clusters:")
        print(" ")
        for i, item in enumerate(self.DataMemberList):
            print(i, item)
        print("**************************")

    def showMaxValues(self):
        print("**************************")
        print("Max Values:")
        print(" ")
        for item in self.MaxValList:
            print(item)
        print("**************************")
    
    def showPythonDAN(self):
        for i, item in enumerate(self.pythonDANList):
            print(len(self.pythonDANList))
            print("**************************")
            print(f"Data Member {i}:")
            print(" ")
            print(item)
            print(" ")
            print(" ")
            print("**************************")

    def showMAXSUBCount(self):
        if self.MAXSUBPython:
            print("**************************")
            print("Max/Sub Counts:")
            print(" ")
            for item in self.MAXSUBListOfListsPython:
                print(item)
            print("**************************")
        else:
            print("No Max/Sub List was created")
    
    def showInputs(self):
        print("Inputs:")
        for input1 in range(len(self.HolderListDFrame[0])):
            print(self.HolderListDFrame[0][input1], self.HolderListDFrame[1][input1])

    def getMaxValue(self):
        return self.MaxValList[0][1]

    def getCategoryMaxValue(self, category, allMaxValues=False, randomMaxValue=True, firstMaxValue=False):
        if self.printStatements:
            print("Getting Category Max Value")
        self.category = category
        self.allMaxValues = allMaxValues
        self.randomMaxValue = randomMaxValue
        self.firstMaxValue = firstMaxValue

        maxVal = self.getMaxValue()
        search = False
        maxValListInCategory = []
        for item in self.MaxValList:
            if item[0] == category:
                search = True
            elif search and item[1] == maxVal:
                maxValListInCategory.append(item)
            elif search and item[1] == 'THISISTHEENDOFTHECATEGORY':
                break
        if allMaxValues:
            finalList = []
            for item in maxValListInCategory:
                finalList.append([category, item[0]])
            if self.printStatements:
                print("Category Max Values Retrieved!")
            return finalList
        elif randomMaxValue:
            holder45 = rn.choice(maxValListInCategory).copy()
            holder45.insert(0, category)
            if self.printStatements:
                print("Category Max Value Retrieved!")
            return holder45[0:2]
        elif firstMaxValue:
            holder46 = maxValListInCategory[0].copy()
            holder46.insert(0, category)
            if self.printStatements:
                print("Category Max Value Retrieved!")
            return holder46[0:2]
        
    def getCategoryMaxValuexMAXSUBCount(self, category):
        if self.printStatements:
            print("Getting Max Max Value x Max Sub")
        self.category = category

        search = False
        MAXSUBxValueList = []
        self.MaxValList[-1][1] = "THISISTHEENDOFTHECATEGORY"
        for index in range(len(self.MaxValList)):
            if self.BigList[index] == category and self.BigList[index - 1] == ' ':
                search = True
            elif search and (isinstance(self.MaxValList[index][1], int) or isinstance(self.MaxValList[index][1], float)):
                for sublist in self.MAXSUBListOfListsPython[1:][::-1]:
                    if sublist[index + 1] != 0 and sublist[index + 1] != None:
                        MAXSUBxValueList.append([category, self.BigList[index], round(self.MaxValList[index][1]*sublist[index + 1], 6)])
                        break
            elif search and self.MaxValList[index][1] == 'THISISTHEENDOFTHECATEGORY':
                holder = []
                for thing in MAXSUBxValueList:
                    holder.append(thing[2])
                maxNum = max(holder)
                finalList = []
                for i, maxNums in enumerate(MAXSUBxValueList):
                    if maxNums[2] == maxNum:
                        finalList.append(maxNums[0:2])
                break
        chosenList = rn.choice(finalList)
        if self.printStatements:
            print("Max Max Value x Max Sub Retrieved")
        return chosenList[0:2]
    
    def getCategoryMAXSUBCountAggregateTotal(self, category):
        if self.printStatements:
            print("Getting Max aggregate of MAX/SUB Count category...")
        self.category = category

        search = False
        MAXSUBxValueList = []
        self.MaxValList[-1][1] = "THISISTHEENDOFTHECATEGORY"
        for index in range(len(self.BigList)):
            if self.BigList[index] == category and self.BigList[index - 1] == ' ':
                search = True
            elif search and (isinstance(self.MaxValList[index][1], int) or isinstance(self.MaxValList[index][1], float)):
                holder12 = []
                for sublist in self.MAXSUBListOfListsPython[1:]:
                    if sublist[index + 1] != None:
                        holder12.append(sublist[0]*sublist[index + 1])
                MAXSUBxValueList.append([category, self.BigList[index], round(sum(holder12), 6)])
            elif (search and self.MaxValList[index][1] == 'THISISTHEENDOFTHECATEGORY'):
                holder = []
                for thing in MAXSUBxValueList:
                    holder.append(thing[2])
                maxNum = max(holder)
                finalList = []
                for i, maxNums in enumerate(MAXSUBxValueList):
                    if maxNums[2] == maxNum:
                        finalList.append(maxNums[0:2])
                break
        chosenList = rn.choice(finalList)
        if self.printStatements:
            print("Max aggregate of MAX/SUB Count retrieved!")
        return chosenList[0:2]
    
    def getContinuousCategoryFunctionApproximationExponentially(self, category, base):
        if self.printStatements:
            print("Getting continuous approximation of category...")
        self.category = category
        self.base = base

        search = False
        MAXSUBxValueList = []
        self.MaxValList[-1][1] = "THISISTHEENDOFTHECATEGORY"
        for index in range(len(self.BigList)):
            if self.BigList[index] == category and self.BigList[index - 1] == ' ':
                search = True
            elif search and (isinstance(self.MaxValList[index][1], int) or isinstance(self.MaxValList[index][1], float)):
                holder12 = []
                for x, sublist in enumerate(self.MAXSUBListOfListsPython[1:]):
                    if sublist[index + 1] != None:
                        holder12.append(sublist[index + 1] * (base ** x))
                MAXSUBxValueList.append([category, self.BigList[index], round(sum(holder12), 6)])
            elif (search and self.MaxValList[index][1] == 'THISISTHEENDOFTHECATEGORY'):
                holder = []
                for thing in MAXSUBxValueList:
                    holder.append(thing[2])
                maxNum = max(holder)
                finalList = []
                for i, maxNums in enumerate(MAXSUBxValueList):
                    if maxNums[2] == maxNum:
                        finalList.append(maxNums[0:2])
                break
        chosenList = rn.choice(finalList)
        if self.printStatements:
            print("Max aggregate of MAX/SUB Count retrieved!")
        return chosenList[0:2]

    def addCluster(self, cluster):
        if self.printStatements:
            print("Adding Cluster...")
        replacementDAN = DAN(
            type=self.type, 
            excelDAN=self.excelDAN,
            pythonDAN=self.pythonDAN,
            MAXSUBPython=self.MAXSUBPython,
            orientation=self.orientation,
            inputFormatting=self.inputFormatting,
            newWorkbook=self.newWorkbook,   
            design=self.design, 
            ListOfLists=self.ListOfLists,
            originalWorkbook=self.originalWorkbook,
            dataSheet=self.dataSheet,
            categoryOrderPreservation=self.categoryOrderPreservation,
            numericalAndAlphabeticalPreservation=self.numericalAndAlphabeticalPreservation,
            allInputCategories=self.allInputCategories,
            desiredModifications=self.desiredModifications,
            categoryNames=self.categoryNames, 
            printStatements=self.printStatements
        )
        self.ListOfLists.append(cluster)
        self.DataMemberList.append(cluster)
        replacementDAN.ListOfLists = self.ListOfLists
        # for i in range(len(self.BigList)):
        #     print(self.BigList[i], self.pythonDANList[-1][i])
        it = replacementDAN.make()
        if self.printStatements:
            print("Cluster added!")
        return it
        

    def changeSetting(self, changeSettingTo):
        self.changeSettingTo = changeSettingTo

        if self.type == "static":
            raise ValueError("This method is not applicable to a static DAN")
        if self.type == "temporal":
            print("?")

    def addData(self, additionalDataWorkbook, additionalDataWorkbooksheet):
        self.additionalDataWorkbook = additionalDataWorkbook
        self.additionalDataWorkbooksheet = additionalDataWorkbooksheet

        if self.type == "static":
            raise ValueError("This method is not applicable to a static DAN, consider adding data to existing excel sheet/ListOfLists and use .make()")
        if self.type == "temporal":
            print("?")
